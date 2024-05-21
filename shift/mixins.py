import calendar
from collections import deque
import datetime
import itertools
from django import forms
from user.models import CustomUser
import pandas as pd
import numpy as np
from .models import Schedule
from django.shortcuts import redirect, render, get_object_or_404
from django.shortcuts import render
from bs4 import BeautifulSoup
import urllib.request
import re
#from openpyxl.writer.excel import save_virtual_workbook
import openpyxl as px
from openpyxl import worksheet
from openpyxl.utils import range_boundaries
import locale
from openpyxl.styles.alignment import Alignment


class BaseCalendarMixin:
    """カレンダー関連Mixinの、基底クラス"""
    first_weekday = 4  # 0は月曜から、1は火曜から。6なら日曜日からになります。お望みなら、継承したビューで指定してください。
    week_names = ['月', '火', '水', '木', '金', '土', '日']  # これは、月曜日から書くことを想定します。['Mon', 'Tue'...

    def setup_calendar(self):
        self._calendar = calendar.Calendar(self.first_weekday)

    def get_week_names(self):
        """first_weekday(最初に表示される曜日)にあわせて、week_namesをシフトする"""
        week_names = deque(self.week_names)
        week_names.rotate(-self.first_weekday)  # リスト内の要素を右に1つずつ移動...なんてときは、dequeを使うと中々面白いです
        return week_names

class EDWmixin(BaseCalendarMixin):
    """週間カレンダーの機能を提供するMixin"""

    def get_previous_week(self, date):
        """前月を返す"""
        if date.month == 1 and date.day ==16:
            return date.replace(year=date.year-1, month=12,day=16)
        if date.month == 1 and date.day ==1:
            return date.replace(year=date.year-1, month=12,day=1)

        if date.day == 16:
            return date.replace(month =date.month -1 ,day = 16)
        if date.day == 1:
            return date.replace(month =date.month -1 ,day = 1)

    def get_week_days(self):
        """その週の日を全て返す"""
        month = self.kwargs.get('month')
        year = self.kwargs.get('year')
        day = self.kwargs.get('day')
        if month and year and day:
            date = datetime.date(year=int(year), month=int(month), day=int(day))

            if date.month ==12 and date.day >10:
                month = 1
                year = date.year +1
                dtm = calendar.monthrange(year,month)[1]
                date = datetime.date(year = int(year),month=int(month),day = int(1))
                dtlist = [date + datetime.timedelta(days =day) for day in range(0,15)]
                return dtlist
            # if date.month ==1  and  date.day ==1:
            #     month =12
            #     year = date.year-1
            #     dtm = calendar.monthrange(year,month)[1]
            #     date = datetime.date(year = int(year),month=int(month),day = int(15))
            #     dtlist = [date + datetime.timedelta(days =day) for day in range(1,dtm-14)]
            #     return dtlist

            if date.month != 12 and date.day < 21 and date.day > 4:
                dtm = calendar.monthrange(year,month)[1]
                date = datetime.date(year = int(year),month=int(month+1),day = int(1))
                dtlist = [date + datetime.timedelta(days =day) for day in range(0,15)]
                return dtlist
            if date.day < 6:
                dtm = calendar.monthrange(year,month)[1]
                date = datetime.date(year = int(year),month=int(month),day = int(15))
                dtlist = [date + datetime.timedelta(days =day) for day in range(1,dtm-14)]
                return dtlist
            if date.day > 20:
                month =month-1
                dtm = calendar.monthrange(year,month)[1]
                date = datetime.date(year = int(year),month=int(month),day = int(15))
                dtlist = [date + datetime.timedelta(days =day) for day in range(1,dtm-14)]
                return dtlist

        else:
            date = datetime.date.today()
            year =int(date.year)
            month=int(date.month)
            day=int(date.day)

            if date.month ==12 and date.day < 21 and date.day > 5:
                month = 1
                year = date.year +1
                dtm = calendar.monthrange(year,month)[1]
                date = datetime.date(year = int(year),month=int(month),day = int(1))
                dtlist = [date + datetime.timedelta(days =day) for day in range(0,15)]
                return dtlist
            
            if date.day-4 < 21 and date.day > 5:
                date = datetime.date(year = int(year),month=int(month+1),day = int(1))
                dtlist = [date + datetime.timedelta(days =day) for day in range(0,15)]
                return dtlist
            if date.day-4 < 6:
                dtm = calendar.monthrange(year,month)[1]
                date = datetime.date(year = int(year),month=int(month),day = int(15))
                dtlist = [date + datetime.timedelta(days =day) for day in range(1,dtm-14)]
                return dtlist
            if date.day-4 > 20:
                month = month + 1
                dtm = calendar.monthrange(year,month)[1]
                date = datetime.date(year = int(year),month=int(month),day = int(15))
                dtlist = [date + datetime.timedelta(days =day) for day in range(1,dtm-14)]
                return dtlist
            


    def name(self):
        weekname=['月','火','水','木','金','土','日']
        name=[]
        week=[]
        day = self.get_week_days()
        for l in day:
            name.append(l.weekday())
        for nn in name:
            week.append(weekname[nn])
        return week


    def get_week_calendar(self):
        """週間カレンダー情報の入った辞書を返す"""
        self.setup_calendar()
        days = self.get_week_days()
        first = days[0]
        last = days[-1]
        ago = self.get_previous_week(first)
        calendar_data = {
            'now': datetime.date.today(),
            'week_days': days,
            'week_previous': ago,
            'week_next': first ,
            'week_names': self.name(),
            'week_first': first,
            'week_last': last,
        }
        return calendar_data

class Shift_OutputMixin(EDWmixin):

    def get_week_schedules(self, start, end, days):
        
        job_pk = self.kwargs['job_pk']
        user= CustomUser.objects.filter(job=job_pk)
       
        lookup = {
            '{}__range'.format(self.date_field): (start, end),
                 
        }
        queryset = self.model.objects.filter(**lookup).order_by('user__user_no')
        days = {day: [] for day in days}   
        df = pd.DataFrame(days)
        b=[ a for a in user]
        # df.loc["必要人数"]=0
        # df.loc["過不足"]=0
        a=1
        for schedule in queryset:
            if schedule.user in b:
                date= schedule.date
                start_time=schedule.get_start_time_display()
                end_time = schedule.get_end_time_display()
                time = start_time+'-'+end_time
                if time =='-':
                    time=None
                if a == 1:
                    user=schedule.user.user_name
                    ddf =pd.DataFrame({date:time},index =[user])
                    df = pd.concat([df,ddf],axis=0)
                    df.fillna(" ", inplace=True)
                    a = 2
                    
                elif user != schedule.user.user_name: 
                    user=schedule.user.user_name
                    ddf =pd.DataFrame({date:time},index =[user])
                    df = pd.concat([df,ddf],axis=0)
                    df.fillna(" ", inplace=True)
                    
                else:    
                    user=schedule.user.user_name
                    df[date]= df[date].astype(str)
                    df.at[user,date] =time
                    df.fillna(" ", inplace=True)
            else:
                pass
                   
               
        df.fillna(" ", inplace=True)    
        return df

    def get_week_calendar(self):
        calendar_context = super().get_week_calendar()
        calendar_context['df'] = self.get_week_schedules(
            calendar_context['week_first'],
            calendar_context['week_last'],
            calendar_context['week_days']
        )
        return calendar_context

class Shift_ConfirmationMixin(EDWmixin):
    """スケジュール付きの、週間カレンダーを提供するMixin"""

    def get_week_schedules(self, start, end, days):
        """それぞれの日とスケジュールを返す"""
        lookup = {
            # '例えば、date__range: (1日, 31日)'を動的に作る
            '{}__range'.format(self.date_field): (start, end),
            'user__pk': self.kwargs.get('user_pk'),
        }
        # 例えば、Schedule.objects.filter(date__range=(1日, 31日)) になる
        queryset = self.model.objects.filter(**lookup)

        # {1日のdatetime: 1日のスケジュール全て, 2日のdatetime: 2日の全て...}のような辞書を作る
        day_schedules = {day: [] for day in days}
        for schedule in queryset:
            schedule_date = getattr(schedule, self.date_field)
            day_schedules[schedule_date].append(schedule)
        return day_schedules
    
    def get_month_calendar(self):
            calendar_context = super().get_week_calendar()
            month_days = calendar_context['week_days']
            month_first = month_days[0]
            month_last = month_days[-1]
            calendar_context['days']=month_days
            calendar_context['week_day_schedules'] = self.get_week_schedules(
            calendar_context['week_first'],
            calendar_context['week_last'],
            calendar_context['week_days']
        )
            
            return calendar_context



class SubmissionMixin(EDWmixin):
    """スケジュール付きの、月間カレンダーを提供するMixin"""

    def get_month_forms(self, start, end, days):
        """それぞれの日と紐づくフォームを作成する"""
        lookup = {
            '{}__range'.format(self.date_field): (start, end),
            'user__pk': self.kwargs.get('user_pk'),
            
        }
        queryset = self.model.objects.filter(**lookup)
        days_count = len(days)
        FormClass = forms.modelformset_factory(self.model, self.form_class, extra=days_count,max_num=days_count)
        if self.request.method == 'POST':

            formset = self.month_formset = FormClass(self.request.POST)
        else:
            formset = self.month_formset = FormClass(queryset=queryset)
        dates =[]
        for bound_form in formset.initial_forms:
            
            instance = bound_form.instance
            date = getattr(instance, self.date_field)
            dates.append(date)
            if date in days:
                days.remove(date)
        day_forms = {day: [] for day in days }

        for empty_form, (date, empty_list) in zip(formset.extra_forms, day_forms.items()):
            empty_form.initial = {self.date_field: date}
            empty_list.append(empty_form)

        for bound_form in formset.initial_forms:
            instance = bound_form.instance
            date = getattr(instance, self.date_field)
            d2 ={date:[]}
            d2[date].append(bound_form)   
            day_forms.update(d2)
        
        day_forms = sorted(day_forms.items())
        day_forms=dict(day_forms)


        return [{key: day_forms[key] for key in itertools.islice(day_forms, 0, days_count)}]

    def get_month_calendar(self):
        calendar_context = super().get_week_calendar()
        month_days = calendar_context['week_days']
        month_first = month_days[0]
        month_last = month_days[-1]
        calendar_context['days']=month_days
        calendar_context['month_day_forms'] = self.get_month_forms(
            month_first,
            month_last,
            month_days
        )
        calendar_context['month_formset'] = self.month_formset
        
        return calendar_context

class Week_CsvMixin(BaseCalendarMixin):
    """週間カレンダーの機能を提供するMixin"""

    def get_previous_week(self, date):
        """前月を返す"""
        if date.month == 1 and date.day ==16:
            return date.replace(year=date.year-1, month=12,day=16)
        if date.month == 1 and date.day ==1:
            return date.replace(year=date.year-1, month=12,day=1)

        if date.day == 16:
            return date.replace(month =date.month -1 ,day = 16)
        if date.day == 1:
            return date.replace(month =date.month -1 ,day = 1)

    def get_week_days(self):
        """その週の日を全て返す"""
        month = self.kwargs.get('month')
        year = self.kwargs.get('year')
        day = self.kwargs.get('day')
        if month and year and day:
            date = datetime.date(year=int(year), month=int(month), day=int(day))

            if date.day < 21 and date.day > 5:
                dtm = calendar.monthrange(year,month)[1]
                date = datetime.date(year = int(year),month=int(month),day=int(15))
                dtlist = [date + datetime.timedelta(days =day) for day in range(1,dtm-14)]
                return dtlist
            if date.day < 6:
                date = datetime.date(year = int(year),month=int(month),day = int(1))
                dtlist = [date + datetime.timedelta(days =day) for day in range(0,15)]
                return dtlist

        else:
            pass

    def name(self):
        weekname=['月','火','水','木','金','土','日']
        name=[]
        week=[]
        day = self.get_week_days()
        for l in day:
            name.append(l.weekday())
        for nn in name:
            week.append(weekname[nn])
        return week


    def get_week_calendar(self):
        """週間カレンダー情報の入った辞書を返す"""
        self.setup_calendar()
        days = self.get_week_days()
        first = days[0]
        last = days[-1]
        ago = self.get_previous_week(first)
        calendar_data = {
            'now': datetime.date.today(),
            'week_days': days,
            'week_previous': ago,
            'week_next': first ,
            'week_names': self.name(),
            'week_first': first,
            'week_last': last,
        }
        return calendar_data        
       

class CsvMixin(Week_CsvMixin):

    def get_week_schedules(self, start, end, days):
        
        
        wb = px.load_workbook('shift/templates/shift/on_template.xlsx')
        sheet = wb['原本']
        day=[]

        
        for a in days:
            b=a.day
            month=a.month
            day.append(b)
        sheet['C2']=str(month)+'月'
        sheet['e2']=day[0]
        sheet['G2']=day[1]
        sheet['I2']=day[2]
        sheet['K2']=day[3]
        sheet['M2']=day[4]
        sheet['O2']=day[5]
        sheet['Q2']=day[6]
        sheet['S2']=day[7]
        sheet['U2']=day[8]
        sheet['W2']=day[9]
        sheet['Y2']=day[10]
        sheet['AA2']=day[11]
        sheet['AC2']=day[12]
        sheet['AE2']=day[13]
        sheet['AG2']=day[14]
        max=5
        if len(day) >= 16:
            max=16
            sheet['AI2']=day[15]



        # shop = get_object_or_404(Shops, pk=self.kwargs['shops_pk'])
        job=self.kwargs['job_pk']
        user= CustomUser.objects.filter(job=job).order_by('user_no')
        b =[]
        users=[]
        kk=1
        for a in user:
            if kk ==1:
                b.append(a)
                name=a.user_name
                start_user={name:['', '', '', '', '', '', '', '', '', '', '', '', '','','','']}
                end_user={name:['', '', '', '', '', '', '', '', '', '', '', '', '','','','']}
                users.append(a)
                kk=2
            else:
                b.append(a)
                name=a.user_name
                start_user.setdefault(name,['', '', '', '', '', '', '', '', '', '', '', '', '','','',''])
                end_user.setdefault(name,['', '', '', '', '', '', '', '', '', '', '', '', '','','',''])
                users.append(a)
        lookup = {
            '{}__range'.format(self.date_field): (start, end),
                 
        }
        queryset = self.model.objects.filter(**lookup).order_by('user__user_no')
        # -----日付のみ出力
        # dd =[]
        # for dday in days:
        #     d =dday.day
        #     dd.append(d)
        # # ^^^^^    
        days = {day: [] for day in days } 
        cf = days
        df = pd.DataFrame(days)
        ff = pd.DataFrame(days)
        i_user=[]
        a=1
        for schedule in queryset:
            
            if schedule.user in b:
                if a == 1:
                    user=schedule.user.user_name
                    date= schedule.date
                    start_time=schedule.get_start_time_display()
                    end_time = schedule.get_end_time_display()
                    fff =pd.DataFrame({date:end_time},index =[user])
                    ff = pd.concat([ff,fff],axis=0)
                    ff.fillna(" ", inplace=True)

                    ddf =pd.DataFrame({date:start_time},index =[user])
                    df = pd.concat([df,ddf],axis=0)
                    df.fillna(" ", inplace=True)
                    ff.fillna(" ", inplace=True)
                    i_user.append(user)
                    a = 2
                    
                if user != schedule.user.user_name: 
                    user=schedule.user.user_name
                    date= schedule.date
                    start_time=schedule.get_start_time_display()
                    end_time = schedule.get_end_time_display()
                    fff =pd.DataFrame({date:end_time},index =[user])
                    ff = pd.concat([ff,fff],axis=0)
                    ff.fillna(" ", inplace=True)


                    ddf =pd.DataFrame({date:start_time},index =[user])
                    df = pd.concat([df,ddf],axis=0)
                    df.fillna(" ", inplace=True)
                    ff.fillna(" ", inplace=True)
                    i_user.append(user)
                    
                else:    
                    user=schedule.user.user_name
                    date= schedule.date
                    start_time=schedule.get_start_time_display()
                    end_time = schedule.get_end_time_display()
                    df[date]= df[date].astype(str)
                    df.at[user,date] =start_time
                    ff.at[user,date]=end_time
                    ff.fillna(" ", inplace=True) 
                    df.fillna(" ", inplace=True) 
        # config= Shop_config_day.objects.filter(**lookup,shops__shop=shop)
        # for shop_config_day in config:
        #     # if shop[0] == shop_config_day.shops:
        #     date = shop_config_day.date
        #     need = shop_config_day.day_need
        #     cf[date] =need
        a=df.values
        b=ff.values
        cc=[]

        for user,aa,bb in zip(i_user,a,b):
            start_user[user]=aa
            end_user[user]=bb

        i=[5,7,9,11,13,15,17,19,21,23,25,27,29,31,33,35]
        j=[6,8,10,12,14,16,18,20,22,24,26,28,30,32,34,36]
        t=15
        ss=1
        op=16
        yo=5
        youbi=self.name()
        for y,config in zip(youbi,cf.values()):
            if config==[]:
                sheet.cell(row=11, column=yo+1, value='0')
            else:
                sheet.cell(row=11, column=yo+1, value=config)
            sheet.cell(row=3, column=yo, value=y)
            
            yo +=2
        users= CustomUser.objects.filter(job=job).order_by('user_no')
        for user,s_time,e_time in zip(users,start_user.values(),end_user.values()):
            t += 1
            user=user.user_name
            sheet.cell(row=op, column=3, value=user)
            op +=1

            for aa,bb,ii,jj in zip(s_time,e_time,i,j):
                if aa=='' or aa==' ':
                    pass
                elif "." in aa:
                    aa=float(aa)
                    sheet.cell(row=t, column=ii, value=aa)
                else:
                    aa=int(aa)
                    sheet.cell(row=t, column=ii, value=aa)
                
                if bb =='' or bb==' ': 
                    pass 
                elif "." in bb :
                    bb=float(bb)
                    sheet.cell(row=t, column=jj, value=bb)
                    
                else:
                    bb=int(bb)
                    sheet.cell(row=t, column=jj, value=bb)
                
        sheet.merge_cells('e2:f2')
        sheet.merge_cells('g2:h2')
        sheet.merge_cells('i2:j2')
        sheet.merge_cells('k2:l2')
        sheet.merge_cells('m2:n2')
        sheet.merge_cells('o2:p2')
        sheet.merge_cells('q2:r2')
        sheet.merge_cells('s2:t2')
        sheet.merge_cells('u2:v2')
        sheet.merge_cells('w2:x2')
        sheet.merge_cells('y2:z2')
        sheet.merge_cells('aa2:ab2')
        sheet.merge_cells('ac2:ad2')
        sheet.merge_cells('ae2:af2')
        sheet.merge_cells('ag2:ah2')
        sheet.merge_cells('ai2:aj2')
        sheet.merge_cells('e3:f3')
        sheet.merge_cells('g3:h3')
        sheet.merge_cells('i3:j3')
        sheet.merge_cells('k3:l3')
        sheet.merge_cells('m3:n3')
        sheet.merge_cells('o3:p3')
        sheet.merge_cells('q3:r3')
        sheet.merge_cells('s3:t3')
        sheet.merge_cells('u3:v3')
        sheet.merge_cells('w3:x3')
        sheet.merge_cells('y3:z3')
        sheet.merge_cells('aa3:ab3')
        sheet.merge_cells('ac3:ad3')
        sheet.merge_cells('ae3:af3')
        sheet.merge_cells('ag3:ah3')
        sheet.merge_cells('ai3:aj3')
        if max==16:
            sheet.merge_cells('ak3:al3')
            sheet.merge_cells('ak2:al2')

        for row in sheet:
            for cell in row:
                if cell.row == 2 or cell.row==3:
                    cell.alignment = Alignment(horizontal = 'center', 
                                                vertical = 'center')

       
        return wb
      
    def get_week_calendar(self):
        calendar_context = super().get_week_calendar()
        calendar_context['wb'] = self.get_week_schedules(
            calendar_context['week_first'],
            calendar_context['week_last'],
            calendar_context['week_days']
        )

        return calendar_context

